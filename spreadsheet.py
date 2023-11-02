import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Data'
test = open('data.txt')
data = test.readlines()

test2 = open('more_data.txt')
data2 = test2.readlines()

for i in range(1,5):
    cellref=sheet.cell(row=i, column=1)
    cellref.value=data[i-1]

for i in range(1,5):
    cellref=sheet.cell(row=i, column=2)
    cellref.value=data2[i-1]


wb.save("newer_data.xlsx")